"""
Прочитать сохранённый csv-файл из задания №19 и сохранить данные в excel-файл,
кроме возраста – столбец с этими данными не нужен.
К заданию прикреплён пример как должно выглядеть содержания итогового файла.
"""
import csv
import openpyxl

#читаю csv-файл
with open('../file_csv1.csv', encoding='utf-8') as r_file:
    file_reader = csv.reader(r_file, delimiter=',')

    counter = 0
    data = []
    for row in file_reader:
        if counter == 0:
            name_of_column = row
        else:
            data.append(row)
        counter += 1
print(name_of_column)
print(data)

#готую рядки для кінцевого файла екскль щоб було як у завданні
row0 = ['']
index = 0
count = 1
while index < len(data):
    row0.append('person' + str(count))
    index += 1
    count += 1
print(row0)


row1_ids = []
row1_ids.append(name_of_column[0])
index = 0
while index < len(data):
    row1_ids.append(data[index][0])
    index += 1
print(row1_ids)

row2_names = []
row2_names.append(name_of_column[1])
index = 0
while index < len(data):
    row2_names.append(data[index][1])
    index += 1
print(row2_names)

row_notuse_age = []
row_notuse_age.append(name_of_column[2])
index = 0
while index < len(data):
    row_notuse_age.append(data[index][2])
    index += 1
print(row_notuse_age)

row3_phone = []
row3_phone.append(name_of_column[3])
index = 0
while index < len(data):
    row3_phone.append(data[index][3])
    index += 1
print(row3_phone)

#створюю ексель файл і записую у сам ексель файл
wb = openpyxl.Workbook()
print(wb)
print(wb.sheetnames)

sheet = wb['Sheet']
print(sheet)

for row_index, row in enumerate((row0, row1_ids, row2_names, row3_phone)):
    for col_index, value in enumerate(row):
        cell = sheet.cell(row=row_index+1, column=col_index+1)
        cell.value = value

wb.save('file_excel.xlsx')
